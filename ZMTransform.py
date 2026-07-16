#!/usr/bin/env python3
"""
Zinc Project - Measurement Analysis

A project for zinc measurement and analysis.
"""

import sys, datetime, os, re, time, configparser, statistics, threading
import tkinter as tk
from tkinter import scrolledtext, filedialog, ttk
from unittest import result
from pathlib import Path
import messagebox as mb
import glob
from typing import Any
import openpyxl
from ctypes import windll
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

_message: str = "Zinc Project - Measurement Analysis"
_data_ora: str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

try: 
    windll.shcore.SetProcessDpiAwareness(1)  # Imposta la consapevolezza DPI per migliorare la resa grafica su schermi ad alta risoluzione
except Exception as e:
    print(f"Errore durante l'impostazione della consapevolezza DPI: {e}")

### Main Application Window ###
root = tk.Tk()
root.title("Zinc Project")  
root.geometry("1000x800")

# --- Area di Log --- #
log_area = scrolledtext.ScrolledText(root, wrap="word") # in questo modo l'area di log avrà una barra di scorrimento verticale
log_area.pack(expand=True, fill="both", padx=10, pady=10)
# ------------------- #

def dict_to_pdf(data: dict, output_file: str | Path) -> None:
    """Create a PDF file with a table of CoilID and corresponding file paths."""
    doc = SimpleDocTemplate(output_file)
    elements = []

    # Create a table with headers
    table_data = [["CoilID", "DateTime", "File Path"]]
    for (coil_id, datetime_value), file_path in data.items():
        table_data.append([coil_id, datetime_value, str(os.path.basename(file_path))])

    # Create a Table object
    table = Table(table_data)
    # Add some styling to the table
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), '#d3d3d3'),  # Header background color
        ('TEXTCOLOR', (0, 0), (-1, 0), '#000000'),   # Header text color
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),       # Center align all cells
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header font
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),      # Header padding
        ('GRID', (0, 0), (-1, -1), 1, '#000000'),     # Grid lines
    ])
    table.setStyle(style)

    elements.append(table)
    doc.build(elements)

def comuni_to_pdf(data: dict, output_file: str | Path) -> None:
    """Create a PDF file with common CoilID and corresponding BS/TS files."""
    
    doc = SimpleDocTemplate(output_file)
    elements = []

    table_data = [["CoilID", "DateTime", "BS File", "TS File"]]

    for (coil_id, datetime_value), files in data.items():
        table_data.append([
            coil_id,
            datetime_value,
            str(os.path.basename(files["BS"])),
            str(os.path.basename(files["TS"]))
        ])

    table = Table(table_data)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), '#d3d3d3'),
        ('TEXTCOLOR', (0, 0), (-1, 0), '#000000'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, '#000000'),
    ])

    table.setStyle(style)

    elements.append(table)
    doc.build(elements)

#---- FUNZIONE PER SCRIVERE NELL'AREA DI LOG ----#
def write_log(message:str): 
    # Scrive un messaggio nell'area di log
    log_area.insert(tk.END, message + "\n")
    log_area.see(tk.END)
    root.update_idletasks()
#-----------------------------------------------#

#---- DIZIONARIO COILID + DATETIME --> EXCEL ----#
def create_dict(data: list[Any]) -> dict:
    files = {}
    for file in data:
        wb = openpyxl.load_workbook(file)
        values = wb["Values"]
        #Il valore del CoilID non è detto che sia sempre nella stessa posizione
        for row in values.iter_rows(min_col=1, max_col=1): 
            cella = row[0]
            if cella.value == "CoilID":
                coil = values.cell(row = cella.row, column= cella.column + 1).value
            if cella.value == "DateTime":
                DateTime = values.cell(row = cella.row, column= cella.column +1).value
        files[coil, DateTime] = file
        wb.close()
    return files
#------------------------------------------------#


#----FUNZIONE DI SALVATAGGIO RISULTATO EXCEL----#
def user_save(result_file: str, result_excel) -> None: 
    result_file = filedialog.asksaveasfilename(title="Salva il file con nome",initialfile="Misurazioni_Zinco.xlsx", defaultextension=".xlsx", filetypes=[("File Excel", "*.xlsx")])
    try: 
        if result_file:
            result_excel.save(result_file)
            ordinamento(result_file)
            write_log("ELABORAZIONE TERMINATA")
    except PermissionError as e:
        write_log(f"Errore di permesso durante il salvataggio del file: {e}")
        mb.showerror("Errore di permesso", f"Non è possibile salvare il file. Controlla se il file è aperto in un'altra applicazione o se hai i permessi necessari. \nDettagli: {e}")
    except Exception as e:
        write_log(f"Errore durante il salvataggio del file: {e}")
#-----------------------------------------------#

#----ORDINAMENTO DI UN FILE EXCEL SECONDO LE PRIME DUE COLONNE ----#
def ordinamento(excel: str) -> None:
    wb = openpyxl.load_workbook(excel)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=True):
        data.append(list(row))
    data.sort(key=lambda x: (x[0],x[1]))
    ws.delete_rows(2, ws.max_row)
    for riga in data:
        ws.append(riga)
    wb.save(excel)
#------------------------------------------------------------------#

#----STRUTTURA DATI----#
def create_data_structures_comuni(data_BS: dict, data_TS: dict) -> dict:
    comuni = {}
    non_comuni_BS = {}
    non_comuni_TS = {}

    for chiave in data_BS:
        if chiave in data_TS:
            comuni[chiave] = {
                "BS": data_BS[chiave],
                "TS": data_TS[chiave]
            }
        else:
            non_comuni_BS[chiave] = data_BS[chiave]

    for chiave in data_TS:
        if chiave not in data_BS:
            non_comuni_TS[chiave] = data_TS[chiave]

    #comuni_to_pdf(comuni, "elementi_comuni.pdf")
    #dict_to_pdf(non_comuni_BS, "non_comuniBS.pdf")
    #dict_to_pdf(non_comuni_TS, "non_comuniTS.pdf")

    return comuni, non_comuni_BS, non_comuni_TS
#-----------------#


#FUNZIONE CHE ESEGUE I CALCOLI 
def calculate_excel(data_BS: list[Any], data_TS: list[Any], result_file: str) -> None:
    print(len(data_BS), len(data_TS))
    #inserisco a dizionario tutti i valori di TS e BS per leggerli UNA SOLA VOLTA
    ts_files = create_dict(data_TS)
    bs_files = create_dict(data_BS)
    bs_files_ordered = dict(sorted(bs_files.items()))
    ts_files_ordered = dict(sorted(ts_files.items()))
    #dict_to_pdf(ts_files_ordered, "TS_CoilID_Files.pdf")
    #dict_to_pdf(bs_files_ordered, "BS_CoilID_Files.pdf")
    comuni, non_comuni_bs, non_comuni_ts = create_data_structures_comuni(bs_files_ordered, ts_files_ordered)

    #se il file excel che contiene i risultati esiste, lo apro
    if os.path.exists(result_file):
        result_excel = openpyxl.load_workbook(result_file) #apro il workbook per i risultati    
        re = result_excel.active
    #altrimenti lo creo
    else: 
        result_excel = openpyxl.Workbook()
        re = result_excel.active
        re.title = "Dati"
        re["A1"] = "CoilID"
        re["B1"] = "DateTime"
        re["C1"] = "BS total length"
        re["D1"] = "BS Nominal"
        re["E1"] = "BS Avg"
        re["F1"] = "BS St.Dev"
        re["G1"] = "BS min"
        re["H1"] = "BS max"
        re["I1"] = "TS total length"
        re["J1"] = "TS Nominal"
        re["K1"] = "TS Avg"
        re["L1"] = "TS St.Dev"
        re["M1"] = "TS min"
        re["N1"] = "TS max"

    #Per ogni chiave del dizionario BS
    try:
        #Ciclo su ogni elemento presente nel file contenente tutti i file comuni.
        for i,(chiave,file) in enumerate(comuni.items(), start = 2): 
            bs_file = file["BS"]
            excel_document_bs  = openpyxl.load_workbook(bs_file)
            values          = excel_document_bs["Values"]
            lengthprofiles  = excel_document_bs["LengthProfiles"]
            Nominal = None
            for row in values.iter_rows(min_col=1, max_col=1): 
                cella_ID = row[0]
                if cella_ID.value == "CoilID":
                    CoilID_BS = values.cell(row = cella_ID.row, column= cella_ID.column + 1).value
                if cella_ID.value == "DateTime":
                    DateTime = values.cell(row = cella_ID.row, column = cella_ID.column + 1).value
                if cella_ID.value == "Coating_BS_Nominal":
                    Nominal = values.cell(row = cella_ID.row, column= cella_ID.column + 1).value
            total_length = lengthprofiles.cell(row=lengthprofiles.max_row, column=1).value
            re.cell(column=1, row=i).value = CoilID_BS
            re.cell(column=2, row=i).value = DateTime
            re.cell(column=4, row=i).value = Nominal

            try: 
                #Calcoli PER BS
                rows_avg    = lengthprofiles.iter_rows(min_row=2, max_row=lengthprofiles.max_row, min_col=2, max_col=2, values_only=True)
                values_avg  = [row[0] for row in rows_avg]
                avg_bs      = sum(values_avg) / len(values_avg)
                massimo_bs  = max(values_avg)
                minimo_bs   = min(values_avg)
                dev_std_bs  = statistics.stdev(values_avg)
                re.cell(column=3, row=i).value = total_length
                re.cell(column=5, row=i).value = avg_bs
                re.cell(column=6, row=i).value = dev_std_bs
                re.cell(column=7, row=i).value = minimo_bs
                re.cell(column=8, row=i).value = massimo_bs
            except Exception as e: 
                write_log(f"{file}: NOT OK BS - {e} \n")
            finally:
                root.update_idletasks()
                excel_document_bs.close()

            ts_file = file["TS"]
            excel_document_ts  = openpyxl.load_workbook(ts_file)
            values              = excel_document_ts["Values"]
            lengthprofiles      = excel_document_ts["LengthProfiles"]
            Nominal = None
            for row in values.iter_rows(min_col=1, max_col=1): 
                cella_ID = row[0]
                if cella_ID.value == "CoilID":
                    CoilID_TS = values.cell(row = cella_ID.row, column= cella_ID.column + 1).value
                if cella_ID.value == "DateTime":
                    DateTime = values.cell(row = cella_ID.row, column = cella_ID.column + 1).value
                if cella_ID.value == "Coating_TS_Nominal":
                    Nominal = values.cell(row = cella_ID.row, column= cella_ID.column + 1).value
            total_length = lengthprofiles.cell(row=lengthprofiles.max_row, column=1).value

            re.cell(column=10, row=i).value = Nominal
            
            try: 
                #Calcoli PER TS
                rows_avg    = lengthprofiles.iter_rows(min_row=2, max_row=lengthprofiles.max_row, min_col=2, max_col=2, values_only=True)
                values_avg  = [row[0] for row in rows_avg]
                avg_ts      = sum(values_avg) / len(values_avg)
                massimo_ts  = max(values_avg)
                minimo_ts   = min(values_avg)
                dev_std_ts  = statistics.stdev(values_avg)
                re.cell(column=9, row=i).value = total_length
                re.cell(column=11, row=i).value = avg_ts
                re.cell(column=12, row=i).value = dev_std_ts
                re.cell(column=13, row=i).value = minimo_ts
                re.cell(column=14, row=i).value = massimo_ts
            except Exception as e: 
                write_log(f"{file}: NOT OK TS - {e} \n")

            finally:
                chargin_bar['value'] = (i / len(comuni)) * 100  # Aggiorna la barra di caricamento
                root.update_idletasks()
                excel_document_ts.close()
    except PermissionError as e:
        write_log(f"Errore di permesso durante il salvataggio del file: {e}")
        mb.showerror("Errore di permesso", f"Non è possibile salvare il file. Controlla se il file è aperto in un'altra applicazione o se hai i permessi necessari. \nDettagli: {e}")
    
    numero_elementi_comuni = len(comuni)
    for i,(chiave,file) in enumerate(non_comuni_bs.items(), start = numero_elementi_comuni + 2):
        excel_document_bs  = openpyxl.load_workbook(file)
        values          = excel_document_bs["Values"]
        lengthprofiles  = excel_document_bs["LengthProfiles"]
        Nominal = None
        for row in values.iter_rows(min_col=1, max_col=1): 
            cella_ID = row[0]
            if cella_ID.value == "CoilID":
                CoilID_BS = values.cell(row = cella_ID.row, column= cella_ID.column + 1).value
            if cella_ID.value == "DateTime":
                DateTime = values.cell(row = cella_ID.row, column = cella_ID.column + 1).value
            if cella_ID.value == "Coating_BS_Nominal":
                Nominal = values.cell(row = cella_ID.row, column= cella_ID.column + 1).value
        total_length = lengthprofiles.cell(row=lengthprofiles.max_row, column=1).value
        re.cell(column=1, row=i).value = CoilID_BS
        re.cell(column=2, row=i).value = DateTime
        re.cell(column=4, row=i).value = Nominal
        try: 
            #Calcoli PER BS
            rows_avg    = lengthprofiles.iter_rows(min_row=2, max_row=lengthprofiles.max_row, min_col=2, max_col=2, values_only=True)
            values_avg  = [row[0] for row in rows_avg]
            avg_bs      = sum(values_avg) / len(values_avg)
            massimo_bs  = max(values_avg)
            minimo_bs   = min(values_avg)
            dev_std_bs  = statistics.stdev(values_avg)
            re.cell(column=3, row=i).value = total_length
            re.cell(column=5, row=i).value = avg_bs
            re.cell(column=6, row=i).value = dev_std_bs
            re.cell(column=7, row=i).value = minimo_bs
            re.cell(column=8, row=i).value = massimo_bs
        except Exception as e: 
            write_log(f"{file}: NOT OK BS - {e} \n")

    numero_elementi_comuni_dopo_bs = numero_elementi_comuni + len(non_comuni_bs) + 2
    for i,(chiave,file) in enumerate(non_comuni_ts.items(), start = numero_elementi_comuni_dopo_bs):
            excel_document_ts  = openpyxl.load_workbook(file)
            values              = excel_document_ts["Values"]
            lengthprofiles      = excel_document_ts["LengthProfiles"]
            Nominal = None
            for row in values.iter_rows(min_col=1, max_col=1): 
                cella_ID = row[0]
                if cella_ID.value == "CoilID":
                    CoilID_TS = values.cell(row = cella_ID.row, column= cella_ID.column + 1).value
                if cella_ID.value == "DateTime":
                    DateTime = values.cell(row = cella_ID.row, column = cella_ID.column + 1).value
                if cella_ID.value == "Coating_TS_Nominal":
                    Nominal = values.cell(row = cella_ID.row, column= cella_ID.column + 1).value
            total_length = lengthprofiles.cell(row=lengthprofiles.max_row, column=1).value

            re.cell(column=1, row=i).value  = CoilID_TS
            re.cell(column=2, row=i).value  = DateTime
            re.cell(column=3, row=i).value  = total_length
            re.cell(column=10, row=i).value = Nominal
            
            try: 
                #Calcoli PER TS
                rows_avg    = lengthprofiles.iter_rows(min_row=2, max_row=lengthprofiles.max_row, min_col=2, max_col=2, values_only=True)
                values_avg  = [row[0] for row in rows_avg]
                avg_ts      = sum(values_avg) / len(values_avg)
                massimo_ts  = max(values_avg)
                minimo_ts   = min(values_avg)
                dev_std_ts  = statistics.stdev(values_avg)
                re.cell(column=9, row=i).value = total_length
                re.cell(column=11, row=i).value = avg_ts
                re.cell(column=12, row=i).value = dev_std_ts
                re.cell(column=13, row=i).value = minimo_ts
                re.cell(column=14, row=i).value = massimo_ts
            except Exception as e: 
                write_log(f"{file}: NOT OK TS - {e} \n")

    user_save(result_file, result_excel)  # Chiamo la funzione per salvare il file con il percorso scelto dall'utente 

def main() -> None:

    ##### Technical sets (X BARRA DI CARICAMENTO) #####
    chargin_bar["value"] = 0 #min value
    chargin_bar["maximum"] = 100 #max value

    cartella = filedialog.askdirectory(title="Select a directory")  #apre una finestra di dialogo
    write_log(f"Cartella selezionata: {cartella}")

    ##### SE LA CARTELLA VIENE SELEZIONATA #####
    if cartella: 
    ##### Selezione della cartella BS e dei file Excel associati #####
        file_excelBS = glob.glob(cartella + "/BS/*.xlsx")  # ottengo una lista di tutti i file Excel nella cartella BS
        file_excelTS = glob.glob(cartella + "/TS/*.xlsx")  # ottengo una lista di tutti i file Excel nella cartella TS 

        numero_fileBS = len(file_excelBS) # ottengo il numero dei file excel dentro alla cartella BS
        numero_fileTS = len(file_excelTS) # ottengo il numero dei file excel dentro alla cartella TS

        write_log(f"Numero di file excel trovati in cartella BS: {numero_fileBS}")
        write_log(f"Numero di file excel trovati in cartella TS: {numero_fileTS}")

        result_file = "Misurazioni Zinco.xlsx"

        if file_excelBS: #se il file excel è stato trovato
            if file_excelTS: 
                #calculate_excel(file_excelBS, file_excelTS, result_file)
                #I calcoli vengono fatti in un thread separato in modo da non bloccare la GUI
                thread = threading.Thread(
                    target=calculate_excel, 
                    args=(file_excelBS, file_excelTS, result_file), 
                    daemon=True,
                    )
                thread.start()
        else: 
            write_log("file excel non trovato in cartella BS")  # Stampa un messaggio se non sono stati trovati file Excel
    else:
        print("No directory selected.")  # Print a message if no directory was selected

    print("Zinc Project Started")


frame_bottoni = tk.Frame(root)
frame_bottoni.pack(pady=10)

#creazione barra di caricamento
chargin_bar = ttk.Progressbar(frame_bottoni, orient = "horizontal", length = 200, mode = "determinate")
chargin_bar.grid(row=0, column=2, padx=5)

btn1 = tk.Button(frame_bottoni, text="Seleziona la cartella", command=main)
btn1.grid(row=0, column=0, sticky="w", padx=5)
btn2 = tk.Button(frame_bottoni, text="Esci", command=root.quit)
btn2.grid(row=0, column=1, sticky="e", padx=5)

root.mainloop()  # Avvia il loop principale dell'applicazione Tkinter